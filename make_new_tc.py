#!/usr/bin/python3.6
#coding:utf-8
from openpyxl import load_workbook
wb = load_workbook("sao_ut.xlsx", read_only=True)
sheet_name_list = wb.get_sheet_names()
print(sheet_name_list )
sheet_num=len(sheet_name_list) 
row0_list = []
row1_list = []
tc_row    = []

my_sheet = wb.get_sheet_by_name(sheet_name_list[0])
rows=my_sheet.max_row
#cols=my_sheet.max_column

for i in range(0,rows):
    print(str(i))
    tc_row.clear()
    for sheet_name in sheet_name_list:
        my_sheet = wb.get_sheet_by_name(sheet_name)
        for cell in list(my_sheet.rows)[i]:
            if i == 0:
                row0_list.append(cell.value) 
            else:
                tc_row.append(cell.value) 
    #print(tc_row)
    if i == 0:
       continue 
    tc_template_name = row0_list[0]
    tc_name         = tc_row[0]
    my_dict=dict(zip(row0_list[1:],tc_row[1:]))
    print(my_dict)

#for row in my_sheet.iter_rows(min_row=1, max_col=3, max_row=2):
#for row in my_sheet.iter_rows('E1:G4'):
#    for cell in row:
#        print(cell.value)
#print(my_sheet.cell('A1'))
Data=my_sheet.cell(row=1,column=1).value
print(Data)
for i, row in enumerate(my_sheet.iter_rows()):
    data = list(row)
    print(str(i))
    for j,dd in enumerate(data): 
        #print(dd.value)
        row1_list.append(dd.value) 
#        print(row1_list)
    #for row in  list(my_sheet.rows)[0]:
    #    for col in row:
    #        print(col.value)
